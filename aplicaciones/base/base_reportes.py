import datetime
from openpyxl import Workbook
from django.http.response import HttpResponse
from django.views.generic import TemplateView
from openpyxl.styles import (
    Alignment,
    Border,
    Font,
    Side)
from string import ascii_letters as abecedario
from aplicaciones.base.utils import *

def cabecera_reporte_excel(_model_name):
    """ TITULO DE REPORTE EN EXCEL.

    Retorna el título del reporte que se está construyendo, en él se indica el nombre del
    modelo del cual pertenece así como la fecha actual que es la fecha en que se crea el
    reporte.

    Parámetros:
    _model_name -- nombre de modelo.

    Variables:
    fecha -- fecha actual.

    """
    fecha = datetime.datetime.now()
    titulo = "REPORTE DE {0} EN FORMATO EXCEL REALIZADO EN LA FECHA: {1}".format(
                                                                _model_name.upper(),
                                                                "%s/%s/%s" % (
                                                                    fecha.day, fecha.month, fecha.year
                                                                ))
    return titulo

def validar_id(valor):
    """ VALIDACIÓN DE CAMPO ID.

    Retorna True o False dependiendo si el campo enviado como parámetro es el
    id, esto se realiza para no mostrar en el reporte el id o clave primaria.

    Parámetros:
    valor -- nombre de campo a evaluar.

    """
    if valor is not 'id':
        return True
    return False

class FormatoReporteExcel():
    """ FORMATO PARA REPORTE EN EXCEL DE CUALQUIER MODELO.

    Esta clase realiza un formato base para generar un reporte en excel para cualquier
    modelo que se desee, en el constructor se definen los parámetros a utilizar, además
    se cuenta con algunos métodos que construyen bloque a bloque el reporte.

    Funciones:
    __init__                        -- función constructora de inicio, en ella se recibe el nombre de la aplicacion
                                       donde se encuentra el modelo a utilizar y el nombre del modelo a utilizar.

        Parámetros:
        _app_name                   -- nombre de aplicacion donde se encuentra el modelo a utilizar.
        _model_name                 -- nombre de tipo Texto de modelo a utilizar.

        Variables:
        _app_name                   -- nombre de aplicacion donde se encuentra el modelo a utilizar.
        _model_name                 -- nombre de modelo a utilizar.
        _modelo                     -- modelo en cuestión a utilizarse, se obtiene por medio de una función
                                       que busca en todo el proyecto el modelo con los parámetros: _app_name
                                       y _model_name.
        _nombres_atributos_modelos  -- lista  con los nombres de todos los atributos que
                                       contiene el modelo a excepción de la clave primaria.
        _consulta                   -- consulta que contiene todos los registros del modelo en cuestión.
        _titulo_cabecera            -- titulo de reporte en excel creado a través de una funcipón que incluye
                                       el nombre del modelo y la fecha de creacion del reporte.
        _libro_trabajo              -- libro de trabajo actual en el que se creará el reporte, es una instancia
                                       de Workbook.
        _hoja_trabajo               -- hoja de trabajo inicial donde se pintará el reporte.

    cabecera_tabla_reporte_excel    -- función que recibe opcionalmente como parámetro la dimensión o altura de las
                                       filas y columnas, y genera todo el encabezado de la tabla en excel.
                                       Por defecto se empezara a pintar la cabecera de la tabla en la fila número 3
                                       y el título desde la celda B1.
                                       La cabecera de la tabla empieza en la letra A y su control de calcula automaticamente
                                       por el parámetro cont, se pinta cada encabezado tomando los valores guardados en la
                                       lista _nombres_atributos_modelos.

        Parámetros:
        dimension_fila              -- parámetro opcional definido por defecto en 15 que indica la altura de las filas.
        dimension_columna           -- parámetro opcional definido por defecto en 25 que indica la altura de las columnas.

        Variables:
        letra_cabecera              -- letra formada automaticamente y que indica el final de la celda que se uniran o merguearán
                                       para ubicar correctamente el titulo de la misma.

    pintar_valores_excel            -- función que pinta los valores correspondientes a la consulta del modelo en cuestion, guardada en la
                                       variable _consulta, se recorre el diccionario y se procede a validar que el campo no corresponda al
                                       atributo id ya que esta será la clave primaria y no se desea renderizar.
                                       Se hace una validación cuando el tipo de dato es un Boolean para escribir True o False según corresponda.

    """
    def __init__(self,_app_name,_model_name):
        self._app_name = _app_name
        self._model_name = _model_name
        self._modelo = obtener_modelo(_app_name,_model_name)
        self._nombres_atributos_modelos = obtener_nombres_atributos_modelos(self._modelo)
        self._consulta = obtener_consulta(self._modelo)
        self._titulo_cabecera = cabecera_reporte_excel(self._model_name)
        self._libro_trabajo = Workbook()
        self._hoja_trabajo = self._libro_trabajo.active

    def cabecera_tabla_reporte_excel(self,dimension_fila = 15,dimension_columna = 25):
        self._hoja_trabajo['B1'].alignment = Alignment(horizontal="center", vertical="center")
        self._hoja_trabajo['B1'].border = Border(left=Side(border_style="thin"), right=Side(border_style="thin"),
                                 top = Side(border_style="thin"), bottom = Side(border_style="thin"))
        self._hoja_trabajo['B1'].font=Font(name='Calibri', size = 12, bold=True)
        self._hoja_trabajo['B1'] = self._titulo_cabecera
        letra_cabecera = '{0}'.format(abecedario[len(self._nombres_atributos_modelos)].upper())
        self._hoja_trabajo.merge_cells('B1:{0}1'.format(letra_cabecera))
        self._hoja_trabajo.row_dimensions[3].height = dimension_fila

        cont = 0
        for atributo in self._nombres_atributos_modelos:
            letra = abecedario[cont].upper()
            self._hoja_trabajo['{0}3'.format(letra)].alignment = Alignment(horizontal="center", vertical="center")
            self._hoja_trabajo['{0}3'.format(letra)].border = Border(left=Side(border_style="thin"), right=Side(border_style="thin"),
                                     top = Side(border_style="thin"), bottom = Side(border_style="thin"))
            self._hoja_trabajo['{0}3'.format(letra)].font=Font(name='Calibri', size = 9, bold=True)
            self._hoja_trabajo['{0}3'.format(letra)] = '{0}'.format(atributo.upper())
            self._hoja_trabajo.column_dimensions["{0}".format(letra)].width = dimension_columna
            cont += 1

    def pintar_valores_excel(self):
        contador_fila = 4
        contador_columnas = 1
        cont = len(self._nombres_atributos_modelos)
        valores = obtener_valor_de_atributos_de_modelo(self._modelo)
        for valor in valores:
            for value in valor:
                if validar_id(value):
                    if cont > 0:
                        self._hoja_trabajo.cell(row=contador_fila, column=contador_columnas).alignment = Alignment(horizontal="center")
                        self._hoja_trabajo.cell(row=contador_fila, column=contador_columnas).border = Border(left=Side(border_style="thin"), right=Side(border_style="thin"),
                                                 top=Side(border_style="thin"), bottom=Side(border_style="thin"))
                        self._hoja_trabajo.cell(row=contador_fila, column=contador_columnas).font = Font(name='Calibri', size=8)
                        if type(valor[value]) is bool:
                            if valor[value] is True:
                                self._hoja_trabajo.cell(row=contador_fila,column=contador_columnas).value = 'True'
                            else:
                                self._hoja_trabajo.cell(row=contador_fila,column=contador_columnas).value = 'False'
                        else:
                            self._hoja_trabajo.cell(row=contador_fila,column=contador_columnas).value = valor[value]
                        contador_columnas += 1
                        cont -= 1
                    if cont is 0:
                        cont = len(self._nombres_atributos_modelos)
            contador_fila += 1
            contador_columnas = 1

    def obtener_reporte_excel(self):
        nombre_archivo ="Reporte {0} en Excel.xlsx".format(self._model_name)
        #Definimos que el tipo de respuesta a devolver es un archivo de microsoft excel
        respuesta = HttpResponse(content_type="application/ms-excel")
        contenido = "attachment; filename={0}".format(nombre_archivo)
        respuesta["Content-Disposition"] = contenido
        self._libro_trabajo.save(respuesta)
        return respuesta

    def construir_reporte(self):
        self.cabecera_tabla_reporte_excel()
        self.pintar_valores_excel()

class ObtenerReporteExcel(TemplateView):
    def get(self,request,_app_name,_model_name,*args,**kwargs):
        reporte = FormatoReporteExcel(_app_name,_model_name)
        reporte.construir_reporte()
        return reporte.obtener_reporte_excel()
